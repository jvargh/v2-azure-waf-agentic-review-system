"""
Excel export with professional formatting using openpyxl
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import Any, Dict, List


def create_styled_excel(assessment_data: Dict[str, Any]) -> BytesIO:
    """Generate professionally formatted Excel workbook for assessment"""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Define styles
    teal_fill = PatternFill(start_color="0B6FA4", end_color="0B6FA4", fill_type="solid")
    white_font = Font(bold=True, size=14, color="FFFFFF")
    white_font_small = Font(bold=True, size=11, color="FFFFFF")
    bold_font = Font(bold=True)
    normal_font = Font(size=10)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    pillar_results = assessment_data.get('pillarResults', [])
    overall_score = assessment_data.get('overallArchitectureScore', 0)
    
    # Calculate metrics
    all_recs = []
    for p in pillar_results:
        all_recs.extend(p.get('recommendations', []))
    high_priority_count = sum(1 for r in all_recs if r.get('priority', '').lower() in ['critical', 'high'])
    
    # Calculate average coverage
    coverages = []
    for p in pillar_results:
        details = p.get('subcategoryDetails', {})
        if details:
            entries = list(details.values())
            with_evidence = sum(1 for d in entries if d.get('found_concepts') or d.get('evidence_found'))
            coverage = round((with_evidence / len(entries)) * 100) if entries else 0
            coverages.append(coverage)
    avg_coverage = round(sum(coverages) / len(coverages)) if coverages else 0
    
    # 1. OVERVIEW SHEET
    ws_overview = wb.create_sheet("Overview")
    
    # Title row
    ws_overview.merge_cells('A1:E1')
    cell = ws_overview['A1']
    cell.value = "Azure Well-Architected Assessment Report"
    cell.font = white_font
    cell.fill = teal_fill
    cell.alignment = center_align
    
    # Info rows
    ws_overview['A3'] = "Assessment Name:"
    ws_overview['B3'] = assessment_data.get('name', 'Unknown')
    ws_overview['A4'] = "Assessment ID:"
    ws_overview['B4'] = assessment_data.get('id', 'Unknown')
    ws_overview['A5'] = "Generated:"
    from datetime import datetime
    ws_overview['B5'] = datetime.now().strftime('%m/%d/%Y, %I:%M:%S %p')
    ws_overview['A6'] = "Overall Architecture Score:"
    ws_overview['B6'] = overall_score
    ws_overview['A7'] = "Average Coverage %:"
    ws_overview['B7'] = avg_coverage
    ws_overview['A8'] = "Total Recommendations:"
    ws_overview['B8'] = len(all_recs)
    ws_overview['A9'] = "High/Critical Recommendations:"
    ws_overview['B9'] = high_priority_count
    
    for row in range(3, 10):
        ws_overview[f'A{row}'].font = bold_font
    
    # Pillar Summary header
    ws_overview.merge_cells('A11:E11')
    cell = ws_overview['A11']
    cell.value = "Pillar Summary"
    cell.font = white_font
    cell.fill = teal_fill
    cell.alignment = center_align
    
    # Column headers
    headers = ['Pillar', 'Score', 'Confidence', 'Coverage %', 'Gaps']
    for col_idx, header in enumerate(headers, 1):
        cell = ws_overview.cell(12, col_idx)
        cell.value = header
        cell.font = white_font_small
        cell.fill = teal_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # Data rows
    for row_idx, pillar in enumerate(pillar_results, 13):
        score = pillar.get('overallScore') or pillar.get('score', 0)
        
        # Calculate coverage and gaps
        details = pillar.get('subcategoryDetails', {})
        coverage = ''
        gaps = ''
        if details:
            entries = list(details.values())
            with_evidence = sum(1 for d in entries if d.get('found_concepts') or d.get('evidence_found'))
            coverage = round((with_evidence / len(entries)) * 100) if entries else 0
            total_missing = sum(len(d.get('missing_concepts', [])) for d in entries)
            gaps = total_missing
        
        ws_overview.cell(row_idx, 1).value = pillar.get('pillar', '')
        ws_overview.cell(row_idx, 2).value = score
        ws_overview.cell(row_idx, 3).value = pillar.get('confidence', '')
        ws_overview.cell(row_idx, 4).value = coverage
        ws_overview.cell(row_idx, 5).value = gaps
        
        for col in range(1, 6):
            cell = ws_overview.cell(row_idx, col)
            cell.border = thin_border
            cell.alignment = left_align
    
    # Column widths
    ws_overview.column_dimensions['A'].width = 32
    ws_overview.column_dimensions['B'].width = 18
    ws_overview.column_dimensions['C'].width = 16
    ws_overview.column_dimensions['D'].width = 14
    ws_overview.column_dimensions['E'].width = 10
    
    # 2. PILLAR SCORES SHEET
    ws_pillar = wb.create_sheet("Pillar Scores")
    
    # Title
    ws_pillar.merge_cells('A1:E1')
    cell = ws_pillar['A1']
    cell.value = "Pillar Scores Breakdown"
    cell.font = white_font
    cell.fill = teal_fill
    cell.alignment = center_align
    
    # Headers
    headers = ['Pillar', 'Overall Score', 'Confidence', 'Coverage %', 'Negative Mentions']
    for col_idx, header in enumerate(headers, 1):
        cell = ws_pillar.cell(3, col_idx)
        cell.value = header
        cell.font = white_font_small
        cell.fill = teal_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # Data
    for row_idx, pillar in enumerate(pillar_results, 4):
        score = pillar.get('overallScore') or pillar.get('score', 0)
        
        details = pillar.get('subcategoryDetails', {})
        coverage = ''
        gaps = ''
        if details:
            entries = list(details.values())
            with_evidence = sum(1 for d in entries if d.get('found_concepts') or d.get('evidence_found'))
            coverage = round((with_evidence / len(entries)) * 100) if entries else 0
            total_missing = sum(len(d.get('missing_concepts', [])) for d in entries)
            gaps = total_missing
        
        ws_pillar.cell(row_idx, 1).value = pillar.get('pillar', '')
        ws_pillar.cell(row_idx, 2).value = score
        ws_pillar.cell(row_idx, 3).value = pillar.get('confidence', '')
        ws_pillar.cell(row_idx, 4).value = coverage
        ws_pillar.cell(row_idx, 5).value = gaps
        
        for col in range(1, 6):
            cell = ws_pillar.cell(row_idx, col)
            cell.border = thin_border
            cell.alignment = left_align
    
    ws_pillar.column_dimensions['A'].width = 28
    ws_pillar.column_dimensions['B'].width = 16
    ws_pillar.column_dimensions['C'].width = 16
    ws_pillar.column_dimensions['D'].width = 14
    ws_pillar.column_dimensions['E'].width = 18
    
    # 3. CONCEPT COVERAGE SHEET
    ws_concept = wb.create_sheet("Concept Coverage")
    
    ws_concept.merge_cells('A1:F1')
    cell = ws_concept['A1']
    cell.value = "Concept Coverage & Justification"
    cell.font = white_font
    cell.fill = teal_fill
    cell.alignment = center_align
    
    headers = ['Pillar', 'Subcategory', 'Score', 'Found Concepts', 'Missing Concepts', 'Justification']
    for col_idx, header in enumerate(headers, 1):
        cell = ws_concept.cell(3, col_idx)
        cell.value = header
        cell.font = white_font_small
        cell.fill = teal_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    row_idx = 4
    for pillar in pillar_results:
        details = pillar.get('subcategoryDetails', {})
        if not details:
            ws_concept.cell(row_idx, 1).value = pillar.get('pillar', '')
            ws_concept.cell(row_idx, 2).value = "No subcategory details available"
            for col in range(1, 7):
                ws_concept.cell(row_idx, col).border = thin_border
            row_idx += 1
        else:
            for detail in details.values():
                found = ', '.join(detail.get('found_concepts', detail.get('evidence_found', []))) or 'None'
                missing = ', '.join(detail.get('missing_concepts', [])) or 'None'
                
                ws_concept.cell(row_idx, 1).value = pillar.get('pillar', '')
                ws_concept.cell(row_idx, 2).value = detail.get('name', '')
                ws_concept.cell(row_idx, 3).value = detail.get('final_score', '')
                ws_concept.cell(row_idx, 4).value = found
                ws_concept.cell(row_idx, 5).value = missing
                ws_concept.cell(row_idx, 6).value = detail.get('justification_text', '')
                
                for col in range(1, 7):
                    cell = ws_concept.cell(row_idx, col)
                    cell.border = thin_border
                    cell.alignment = left_align
                
                row_idx += 1
    
    ws_concept.column_dimensions['A'].width = 24
    ws_concept.column_dimensions['B'].width = 36
    ws_concept.column_dimensions['C'].width = 8
    ws_concept.column_dimensions['D'].width = 48
    ws_concept.column_dimensions['E'].width = 48
    ws_concept.column_dimensions['F'].width = 80
    
    # 4. RECOMMENDATIONS SHEET
    ws_recs = wb.create_sheet("Recommendations")
    
    ws_recs.merge_cells('A1:H1')
    cell = ws_recs['A1']
    cell.value = "Recommendations by Pillar"
    cell.font = white_font
    cell.fill = teal_fill
    cell.alignment = center_align
    
    headers = ['Pillar', '#', 'Title', 'Priority', 'Effort', 'Impact', 'Recommendation', 'Source']
    for col_idx, header in enumerate(headers, 1):
        cell = ws_recs.cell(3, col_idx)
        cell.value = header
        cell.font = white_font_small
        cell.fill = teal_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    row_idx = 4
    for pillar in pillar_results:
        recs = pillar.get('recommendations', [])
        if not recs:
            ws_recs.cell(row_idx, 1).value = pillar.get('pillar', '')
            ws_recs.cell(row_idx, 3).value = "No recommendations available"
            for col in range(1, 9):
                ws_recs.cell(row_idx, col).border = thin_border
            row_idx += 1
        else:
            # Sort by priority
            priority_order = {'critical': 1, 'high': 2, 'medium': 3, 'low': 4}
            sorted_recs = sorted(recs, key=lambda r: priority_order.get(r.get('priority', '').lower(), 5))
            
            for idx, rec in enumerate(sorted_recs, 1):
                reasoning = rec.get('reasoning') or rec.get('insight') or rec.get('recommendation') or rec.get('details', '')
                impact = rec.get('business_impact') or rec.get('impact', '')
                
                ws_recs.cell(row_idx, 1).value = pillar.get('pillar', '')
                ws_recs.cell(row_idx, 2).value = idx
                ws_recs.cell(row_idx, 3).value = rec.get('title', '')
                ws_recs.cell(row_idx, 4).value = rec.get('priority', '')
                ws_recs.cell(row_idx, 5).value = rec.get('effort', '')
                ws_recs.cell(row_idx, 6).value = impact
                ws_recs.cell(row_idx, 7).value = reasoning
                ws_recs.cell(row_idx, 8).value = rec.get('source', '')
                
                for col in range(1, 9):
                    cell = ws_recs.cell(row_idx, col)
                    cell.border = thin_border
                    cell.alignment = left_align
                
                # Priority color
                priority = rec.get('priority', '').lower()
                if priority in ['critical', 'high', 'medium', 'low']:
                    color_map = {
                        'critical': 'DC3545',
                        'high': 'FD7E14',
                        'medium': 'FFC107',
                        'low': '28A745'
                    }
                    cell = ws_recs.cell(row_idx, 4)
                    cell.fill = PatternFill(start_color=color_map[priority], end_color=color_map[priority], fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                
                row_idx += 1
    
    ws_recs.column_dimensions['A'].width = 20
    ws_recs.column_dimensions['B'].width = 4
    ws_recs.column_dimensions['C'].width = 42
    ws_recs.column_dimensions['D'].width = 12
    ws_recs.column_dimensions['E'].width = 10
    ws_recs.column_dimensions['F'].width = 46
    ws_recs.column_dimensions['G'].width = 80
    ws_recs.column_dimensions['H'].width = 44
    
    # 5. COHESIVE RECOMMENDATIONS (if present)
    cohesive = assessment_data.get('cohesiveRecommendations', [])
    if cohesive:
        ws_coh = wb.create_sheet("Cohesive Recs")
        
        ws_coh.merge_cells('A1:G1')
        cell = ws_coh['A1']
        cell.value = "Cross-Pillar Cohesive Recommendations"
        cell.font = white_font
        cell.fill = teal_fill
        cell.alignment = center_align
        
        headers = ['#', 'Title', 'Priority', 'Effort', 'Impact', 'Recommendation', 'Source']
        for col_idx, header in enumerate(headers, 1):
            cell = ws_coh.cell(3, col_idx)
            cell.value = header
            cell.font = white_font_small
            cell.fill = teal_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        for row_idx, rec in enumerate(cohesive, 4):
            ws_coh.cell(row_idx, 1).value = row_idx - 3
            ws_coh.cell(row_idx, 2).value = rec.get('title', '')
            ws_coh.cell(row_idx, 3).value = rec.get('priority', '')
            ws_coh.cell(row_idx, 4).value = rec.get('effort', '')
            ws_coh.cell(row_idx, 5).value = rec.get('impact') or rec.get('business_impact', '')
            ws_coh.cell(row_idx, 6).value = rec.get('reasoning') or rec.get('recommendation') or rec.get('details', '')
            ws_coh.cell(row_idx, 7).value = rec.get('source', '')
            
            for col in range(1, 8):
                cell = ws_coh.cell(row_idx, col)
                cell.border = thin_border
                cell.alignment = left_align
            
            priority = rec.get('priority', '').lower()
            if priority in ['critical', 'high', 'medium', 'low']:
                color_map = {
                    'critical': 'DC3545',
                    'high': 'FD7E14',
                    'medium': 'FFC107',
                    'low': '28A745'
                }
                cell = ws_coh.cell(row_idx, 3)
                cell.fill = PatternFill(start_color=color_map[priority], end_color=color_map[priority], fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
        
        ws_coh.column_dimensions['A'].width = 4
        ws_coh.column_dimensions['B'].width = 38
        ws_coh.column_dimensions['C'].width = 12
        ws_coh.column_dimensions['D'].width = 10
        ws_coh.column_dimensions['E'].width = 46
        ws_coh.column_dimensions['F'].width = 80
        ws_coh.column_dimensions['G'].width = 44
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
